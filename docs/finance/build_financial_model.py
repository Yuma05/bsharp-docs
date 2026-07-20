#!/usr/bin/env python3
"""Bsharp 財務モデル (PL/CF) の Excel 生成スクリプト。

使い方:
    python3 docs/finance/build_financial_model.py
    → docs/finance/financial-model.xlsx を生成（既存ファイルは上書き）

設計方針:
- xlsx には「値」ではなく「Excel 数式」を書き込む。
  生成後は Excel / Google Sheets 上で「前提」シートのパラメータを
  直接編集すれば PL/CF が再計算される。日常的な感度分析にこのスクリプトの
  再実行は不要。
- このスクリプトを触るのは「構造を変えるとき」だけ
  （行の追加、計算ロジックの変更、期間の延長など）。
- PARAMS / BUCKETS がパラメータの初期値。シート構造は build_* 関数。
- preview() は Excel 式と同じ計算の Python ミラー。生成時に主要数値を
  標準出力して式の妥当性チェックに使う。**計算ロジックを変えたら
  build_pl/build_cf と preview の両方を更新すること。**

見た目の規約（ikka-brand の財務モデルのデザインシステムを踏襲）:
- 青字 = 入力セル（編集してよい） / 黒 = 自動計算 / 緑 = 他シート参照
- セクション行 = 薄グレー背景 + 左の青アクセントボーダー
- 負値は括弧表示、ゼロは "-"。M12|M13 の間に年境界の縦罫線

前提ドキュメント: 同ディレクトリの README.md / 01-revenue-model.md / 02-cost-structure.md
依存: openpyxl (pip3 install openpyxl)
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MONTHS = 18  # PL の月数 (M1..M18)。CF は M0(開業準備月) を先頭に持つ

# ---------------------------------------------------------------------------
# パラメータ初期値 (key, ラベル, 値, 表示形式, 備考)
# 値を変えたいだけなら xlsx 側で編集すればよい。ここは再生成時の初期値。
# ---------------------------------------------------------------------------
PARAMS = [
    ("sec_rev",     "売上サイド", None, None, ""),
    ("scenario",    "シナリオ係数（セッションに乗算）", 1.0, "0.0", "保守0.6 / 標準1.0 / 強気1.4"),
    ("sess_init",   "セッション数 初月(M1)", 400, "#,##0", "知人・コミュニティ告知込み"),
    ("sess_grow",   "セッション数 月次増分", 320, "#,##0", "週3〜4本のリール/ショート+月2〜4本ブログ前提"),
    ("cvr_base",    "CVRベース（オーガニック）", 0.012, "0.00%", "EC相場1%への上振れ仮説"),
    ("cvr_slope",   "CVR月次改善（M4以降/月）", 0.00011, "0.000%", "レビュー・実績蓄積効果。M12で約1.3%"),
    ("boost_m1",    "知人ブースト注文 M1", 6, "#,##0", "開店告知の初期注文"),
    ("boost_m2",    "知人ブースト注文 M2", 3, "#,##0", ""),
    ("repeat_rate", "月次リピート率", 0.03, "0.0%", "累計購入者に対する月次再購入率"),
    ("ad_start",    "広告開始月", 3, "#,##0", "オーディエンス1,000人ゲート。目安M3"),
    ("ad_budget",   "広告費/月（Meta・リタゲ中心）", 10000, "#,##0", "確定方針"),
    ("rt_cpc",      "リタゲCPC", 60, "#,##0", "相場仮説"),
    ("rt_cvr",      "リタゲCVR", 0.03, "0.0%", "相場仮説"),

    ("sec_cost",    "コストサイド", None, None, ""),
    ("kakeritsu",   "仕入掛率", 0.6, "0.0", "全件交渉未確定。最重要感度変数"),
    ("pay_rate",    "決済手数料率", 0.0355, "0.00%", "Shopify Payments Basic 国内カード（確定）"),
    ("loss_rate",   "破損・不良率（商品売上比）", 0.015, "0.0%", "再送コスト込みの引当"),
    ("ship_charge", "顧客負担送料（未到達時・全国一律）", 880, "#,##0", "設計値"),
    ("free_line",   "送料無料ライン", 11000, "#,##0", "高ライン=送料収入最大化。到達率≈0%。詳細は02"),
    ("po_ship",     "仕入送料/発注回", 2000, "#,##0", "仕入れ先条件で置換"),

    ("sec_fix",     "固定費（月次）", None, None, ""),
    ("fix_shopify", "Shopify Basic（年払い月割）", 3650, "#,##0", "現金は年1回¥43,800"),
    ("fix_domain",  "独自ドメイン（月割）", 170, "#,##0", "現金は年1回¥2,000"),
    ("fix_photo",   "撮影ランニング", 5000, "#,##0", "スタジオ2〜3回/年+撮影消耗品の月割"),

    ("sec_cf",      "初期投資・CF設定", None, None, ""),
    ("inv_initial", "初期仕入（原価）", 440000, "#,##0", "縮小・回転型: 主力10SKU×10個+他20SKU×5〜6個"),
    ("init_photo",  "撮影初期（キービジュアル・備品）", 75000, "#,##0", ""),
    ("init_mat",    "梱包資材 初期ロット", 40000, "#,##0", "月次資材費はこの在庫を消化後に現金化"),
    ("init_open",   "開業その他", 10000, "#,##0", "名刺・印鑑等"),
    ("shopify_year","Shopify 年払い額", 43800, "#,##0", "M0とM12に支払"),
    ("domain_year", "ドメイン 年額", 2000, "#,##0", "M0とM12に支払"),
    ("stock_cover", "在庫カバー月数（発注基準）", 2.0, "0.0", "期末在庫≥翌月原価×この月数となるよう発注"),
]

# 特に重要な入力セル（ハイライト表示）
IMPORTANT = {"scenario", "kakeritsu", "free_line", "cvr_base", "sess_grow"}

# 注文構成バケット: (ラベル, 構成比, 注文額, 送料実費, 資材費)
# 送料実費 = ゆうパック東京発・全国人口加重(近55/中30/遠15)・スマホ割-180、二重梱包サイズ推定
BUCKETS = [
    ("1点買い（60〜80サイズ）", 0.50, 3000, 870, 250),
    ("2点買い（80サイズ）",     0.27, 5800, 1082, 320),
    ("3点以上（80〜100サイズ）", 0.23, 8800, 1183, 420),
]

# ---------------------------------------------------------------------------
# デザインシステム（ikka-brand 財務モデル踏襲）
# ---------------------------------------------------------------------------
FN = "Yu Gothic"
C_PRI = "2D3436"   # 基本テキスト
C_SEC = "636E72"   # 補足テキスト
C_INP = "2D6CDF"   # 入力セル（青字）
C_LNK = "0D7C66"   # 他シート参照（緑字）

title_f = Font(name=FN, bold=True, size=14, color=C_PRI)
sub_f = Font(name=FN, size=8, color=C_SEC)
sec_f = Font(name=FN, bold=True, size=10, color=C_PRI)
hdr_f = Font(name=FN, bold=True, size=9, color=C_PRI)
lbl_f = Font(name=FN, size=9, color=C_PRI)
inp_f = Font(name=FN, size=9, color=C_INP)
calc_f = Font(name=FN, size=9, color=C_PRI)
link_f = Font(name=FN, size=9, color=C_LNK)
bold_f = Font(name=FN, bold=True, size=9, color=C_PRI)

hdr_fill = PatternFill("solid", fgColor="F1F2F6")
sec_fill = PatternFill("solid", fgColor="E8ECF0")
hi_fill = PatternFill("solid", fgColor="EBF5FB")

lt_side = Side(style="thin", color="E0E0E0")
md_side = Side(style="thin", color="B0B0B0")
ac_side = Side(style="medium", color=C_INP)
thin_b = Border(bottom=lt_side)
med_b = Border(bottom=md_side)
ctr = Alignment(horizontal="center")
vctr = Alignment(vertical="center")

JPY = '#,##0;(#,##0);"-"'
NUM1 = '#,##0.0;(#,##0.0);"-"'
NUM = '#,##0;(#,##0);"-"'
PCT = "0.0%"

H_TITLE, H_SEC, H_DATA = 28, 22, 18


def sec_row(ws, r, ncols, label):
    """セクション見出し行: 薄グレー背景 + 左の青アクセントボーダー。"""
    cell = ws.cell(r, 1, label)
    cell.font = sec_f
    cell.alignment = vctr
    for i in range(1, ncols + 1):
        ws.cell(r, i).fill = sec_fill
    ws.cell(r, 1).border = Border(left=ac_side)
    ws.row_dimensions[r].height = H_SEC


def year_border(ws, col_letter, r_min, r_max):
    """年境界（M12|M13）の縦罫線。"""
    for r in range(r_min, r_max + 1):
        cell = ws[f"{col_letter}{r}"]
        b = cell.border
        cell.border = Border(left=b.left, right=md_side, top=b.top, bottom=b.bottom)


# ---------------------------------------------------------------------------
# 「前提」シート
# ---------------------------------------------------------------------------
def build_assumptions(wb):
    ws = wb.active
    ws.title = "前提"
    ws.sheet_properties.tabColor = C_INP
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 52
    for l in ("D", "E", "F"):
        ws.column_dimensions[l].width = 11

    ws["A1"] = "Bsharp 財務モデル"
    ws["A1"].font = title_f
    ws["A2"] = "前提条件・入力パラメータ（青字 = 可変入力セル / 黒 = 自動計算）"
    ws["A2"].font = sub_f
    ws.row_dimensions[1].height = H_TITLE
    ws.row_dimensions[2].height = H_DATA

    for i, h in enumerate(["パラメータ", "値", "備考"]):
        cell = ws.cell(3, 1 + i, h)
        cell.font, cell.fill, cell.alignment = hdr_f, hdr_fill, ctr
    ws.row_dimensions[3].height = H_DATA

    rows = {}  # key -> 行番号
    r = 4
    for key, label, value, fmt, note in PARAMS:
        if value is None:  # セクション見出し
            sec_row(ws, r, 6, label)
        else:
            ws.cell(r, 1, label).font = lbl_f
            cell = ws.cell(r, 2, value)
            cell.font = inp_f
            if fmt:
                cell.number_format = fmt
            cell.border = thin_b
            if key in IMPORTANT:
                cell.fill = hi_fill
            ws.cell(r, 3, note).font = sub_f
            ws.row_dimensions[r].height = H_DATA
            rows[key] = r
        r += 1

    # --- 注文構成バケット表 ---
    sec_row(ws, r, 6, "注文構成（送料無料ライン到達は自動判定）")
    r += 1
    for i, h in enumerate(["バケット", "構成比", "注文額", "送料実費", "資材費", "到達フラグ"]):
        cell = ws.cell(r, 1 + i, h)
        cell.font, cell.fill, cell.alignment = hdr_f, hdr_fill, ctr
    ws.row_dimensions[r].height = H_DATA
    bucket_start = r + 1
    for i, (label, share, value, ship, mat) in enumerate(BUCKETS):
        br = bucket_start + i
        ws.cell(br, 1, label).font = lbl_f
        for c_i, (v, fmt) in enumerate(
                [(share, "0%"), (value, NUM), (ship, NUM), (mat, NUM)], start=2):
            cell = ws.cell(br, c_i, v)
            cell.font, cell.number_format, cell.border = inp_f, fmt, thin_b
        flag = ws.cell(br, 6, f"=IF(C{br}>=B{rows['free_line']},1,0)")
        flag.font, flag.border, flag.alignment = calc_f, thin_b, ctr
        ws.row_dimensions[br].height = H_DATA
    bucket_end = bucket_start + len(BUCKETS) - 1
    shares = f"B{bucket_start}:B{bucket_end}"
    values = f"C{bucket_start}:C{bucket_end}"
    ships = f"D{bucket_start}:D{bucket_end}"
    mats = f"E{bucket_start}:E{bucket_end}"
    flags = f"F{bucket_start}:F{bucket_end}"

    # --- 計算セル（ブレンド値。PL/CF はここだけを参照する） ---
    r = bucket_end + 1
    sec_row(ws, r, 6, "計算値（自動・編集不要）")
    derived = [
        ("aov", "AOV（平均注文額・商品のみ）", f"=SUMPRODUCT({shares},{values})", NUM),
        ("ship_income", "送料収入/注文（未到達×顧客負担）",
         f"=SUMPRODUCT({shares},1-{flags})*B{rows['ship_charge']}", NUM),
        ("ship_cost", "発送送料実費/注文", f"=SUMPRODUCT({shares},{ships})", NUM),
        ("mat_cost", "梱包資材/注文", f"=SUMPRODUCT({shares},{mats})", NUM),
        ("reach_rate", "無料ライン到達率", f"=SUMPRODUCT({shares},{flags})", PCT),
        ("contrib_per_order", "貢献利益/注文", None, NUM),  # 下で組み立て
        ("fix_total", "固定費 月計（広告除く）",
         f"=B{rows['fix_shopify']}+B{rows['fix_domain']}+B{rows['fix_photo']}", NUM),
    ]
    for i, (key, label, formula, fmt) in enumerate(derived):
        dr = r + 1 + i
        rows[key] = dr
        ws.cell(dr, 1, label).font = lbl_f
        if formula:
            cell = ws.cell(dr, 2, formula)
            cell.font, cell.number_format, cell.border = calc_f, fmt, thin_b
        ws.row_dimensions[dr].height = H_DATA
    # 貢献利益/注文 = AOV+送料収入 − 原価 − 決済 − 発送 − 資材 − 破損
    a, si = f"B{rows['aov']}", f"B{rows['ship_income']}"
    cpo = ws.cell(rows["contrib_per_order"], 2,
                  f"={a}+{si}-{a}*B{rows['kakeritsu']}-({a}+{si})*B{rows['pay_rate']}"
                  f"-B{rows['ship_cost']}-B{rows['mat_cost']}-{a}*B{rows['loss_rate']}")
    cpo.font, cpo.number_format = bold_f, NUM
    cpo.fill, cpo.border = hi_fill, med_b

    ws.freeze_panes = "B4"
    return rows


def ref(rows, key):
    """前提シートのパラメータ参照式を返す。"""
    return f"前提!$B${rows[key]}"


# ---------------------------------------------------------------------------
# 「PL」シート (月次 M1..M18)
# 行1: タイトル / 行2: 月ラベル / 以降セクション+データ
# 月番号は行として持たず、生成時に数式へ直接埋め込む（M0/M1ヘッダーで判別可能なため）
# ---------------------------------------------------------------------------
PL_SECTIONS = {3: "集客・注文", 13: "売上", 18: "変動費", 26: "利益"}
PL_ROWS = {  # 行番号の一元管理。CF からも参照される
    "sessions": 4, "cvr": 5, "org_orders": 6, "boost_orders": 7, "rt_orders": 8,
    "new_orders": 9, "cum_customers": 10, "repeat_orders": 11, "orders": 12,
    "aov": 14, "product_rev": 15, "ship_rev": 16, "revenue": 17,
    "cogs": 19, "gross": 20, "pay_fee": 21, "ship_cost": 22, "mat_cost": 23,
    "loss": 24, "po_ship": 25,
    "contrib": 27, "contrib_rate": 28, "ad": 29, "fixed": 30,
    "op": 31, "cum_op": 32,
}
PL_BOLD = {"new_orders", "orders", "revenue", "gross", "contrib", "op", "cum_op"}
PL_HI = {"contrib", "op"}
PL_INDENT = {"org_orders", "boost_orders", "rt_orders", "repeat_orders",
             "product_rev", "ship_rev", "pay_fee", "ship_cost", "mat_cost",
             "loss", "po_ship"}


def pl_col(m):
    """PL シートで月 m (1..18) の列文字を返す。M1=B列。"""
    return get_column_letter(m + 1)


def cf_col(m):
    """CF シートで月 m (0..18) の列文字を返す。M0=B列。"""
    return get_column_letter(m + 2)


def month_header(ws, ncols, first_label_col=2, m_start=1):
    """行1タイトル余白 + 行2に月ラベルヘッダーを描く。"""
    ws.row_dimensions[1].height = H_TITLE
    ws.row_dimensions[2].height = H_DATA
    hdr = ws.cell(2, 1, "")
    hdr.fill = hdr_fill
    for i in range(ncols - first_label_col + 1):
        c = first_label_col + i
        cell = ws.cell(2, c, f"M{m_start + i}")
        cell.font, cell.fill, cell.alignment = hdr_f, hdr_fill, ctr
        ws.column_dimensions[get_column_letter(c)].width = 11


def build_pl(wb, rows):
    ws = wb.create_sheet("PL")
    ws.sheet_properties.tabColor = C_LNK
    ws.column_dimensions["A"].width = 30
    ws["A1"] = "月次PL（M1〜M18）"
    ws["A1"].font = title_f
    month_header(ws, MONTHS + 1)

    labels = {
        "sessions": "セッション数", "cvr": "CVR",
        "org_orders": "オーガニック新規注文", "boost_orders": "知人ブースト注文",
        "rt_orders": "リタゲ注文", "new_orders": "新規注文 計",
        "cum_customers": "累計購入者数", "repeat_orders": "リピート注文",
        "orders": "総注文数", "aov": "AOV", "product_rev": "商品売上",
        "ship_rev": "送料収入", "revenue": "売上高 合計", "cogs": "売上原価",
        "gross": "売上総利益", "pay_fee": "決済手数料", "ship_cost": "発送送料",
        "mat_cost": "梱包資材費", "loss": "破損・不良", "po_ship": "仕入送料",
        "contrib": "貢献利益", "contrib_rate": "貢献利益率", "ad": "広告費",
        "fixed": "固定費", "op": "営業利益", "cum_op": "累計営業利益",
    }
    for r, label in PL_SECTIONS.items():
        sec_row(ws, r, MONTHS + 1, label)
    for key, r in PL_ROWS.items():
        name = ("  " + labels[key]) if key in PL_INDENT else labels[key]
        ws.cell(r, 1, name).font = bold_f if key in PL_BOLD else lbl_f
        ws.row_dimensions[r].height = H_DATA

    R = PL_ROWS
    for m in range(1, MONTHS + 1):
        c = pl_col(m)
        p = pl_col(m - 1) if m > 1 else None  # 前月列
        # 月番号 m は生成時に数式へ直接埋め込む。
        # 知人ブーストのように条件が構造（何月か）だけに依存するものは生成時に解決し、
        # 広告開始月のように「前提」シートで編集可能なパラメータに依存する条件は IF を残す。
        f = {
            "sessions": f"=ROUND(({ref(rows,'sess_init')}+{ref(rows,'sess_grow')}"
                        f"*{m - 1})*{ref(rows,'scenario')},0)",
            "cvr": f"={ref(rows,'cvr_base')}+{ref(rows,'cvr_slope')}*{max(0, m - 3)}",
            "org_orders": f"={c}{R['sessions']}*{c}{R['cvr']}",
            "boost_orders": f"={ref(rows,'boost_m1')}" if m == 1
                            else f"={ref(rows,'boost_m2')}" if m == 2 else "=0",
            "rt_orders": f"=IF({m}>={ref(rows,'ad_start')},"
                         f"{ref(rows,'ad_budget')}/{ref(rows,'rt_cpc')}*{ref(rows,'rt_cvr')},0)",
            "new_orders": f"=SUM({c}{R['org_orders']}:{c}{R['rt_orders']})",
            "cum_customers": f"={c}{R['new_orders']}" if m == 1
                             else f"={p}{R['cum_customers']}+{c}{R['new_orders']}",
            "repeat_orders": "=0" if m == 1
                             else f"={p}{R['cum_customers']}*{ref(rows,'repeat_rate')}",
            "orders": f"={c}{R['new_orders']}+{c}{R['repeat_orders']}",
            "aov": f"={ref(rows,'aov')}",
            "product_rev": f"={c}{R['orders']}*{c}{R['aov']}",
            "ship_rev": f"={c}{R['orders']}*{ref(rows,'ship_income')}",
            "revenue": f"={c}{R['product_rev']}+{c}{R['ship_rev']}",
            "cogs": f"={c}{R['product_rev']}*{ref(rows,'kakeritsu')}",
            "gross": f"={c}{R['revenue']}-{c}{R['cogs']}",
            "pay_fee": f"={c}{R['revenue']}*{ref(rows,'pay_rate')}",
            "ship_cost": f"={c}{R['orders']}*{ref(rows,'ship_cost')}",
            "mat_cost": f"={c}{R['orders']}*{ref(rows,'mat_cost')}",
            "loss": f"={c}{R['product_rev']}*{ref(rows,'loss_rate')}",
            "po_ship": f"=CF!{cf_col(m)}{CF_ROWS['po_ship']}",
            "contrib": f"={c}{R['gross']}-{c}{R['pay_fee']}-{c}{R['ship_cost']}"
                       f"-{c}{R['mat_cost']}-{c}{R['loss']}-{c}{R['po_ship']}",
            "contrib_rate": f"=IF({c}{R['revenue']}=0,0,{c}{R['contrib']}/{c}{R['revenue']})",
            "ad": f"=IF({m}>={ref(rows,'ad_start')},{ref(rows,'ad_budget')},0)",
            "fixed": f"={ref(rows,'fix_total')}",
            "op": f"={c}{R['contrib']}-{c}{R['ad']}-{c}{R['fixed']}",
            "cum_op": f"={c}{R['op']}" if m == 1 else f"={p}{R['cum_op']}+{c}{R['op']}",
        }
        for key, formula in f.items():
            cell = ws.cell(R[key], m + 1, formula)
            cell.border = thin_b
            cell.font = link_f if key == "po_ship" else \
                bold_f if key in PL_BOLD else calc_f
            if key == "cvr":
                cell.number_format = "0.00%"
            elif key == "contrib_rate":
                cell.number_format = PCT
            elif key == "sessions":
                cell.number_format = NUM
            elif key in ("org_orders", "boost_orders", "rt_orders", "new_orders",
                         "cum_customers", "repeat_orders", "orders"):
                cell.number_format = NUM1
            else:
                cell.number_format = JPY
            if key in PL_HI:
                cell.fill = hi_fill
            if key in ("new_orders", "orders", "revenue", "op"):
                cell.border = med_b
    year_border(ws, pl_col(12), 2, max(PL_ROWS.values()))
    ws.freeze_panes = "B3"
    return ws


# ---------------------------------------------------------------------------
# 「CF」シート (月次 M0..M18)。仕入は前払い・同月納品、売上入金は同月(Shopify週次入金の近似)
# ---------------------------------------------------------------------------
CF_SECTIONS = {3: "現金収入", 5: "在庫（原価ベース）", 10: "現金支出", 21: "キャッシュフロー"}
CF_ROWS = {
    "cash_in": 4,
    "stock_open": 6, "purchase": 7, "cogs_use": 8, "stock_close": 9,
    "pay_purchase": 11, "po_ship": 12, "var_cash": 13,
    "mat_cum": 14, "mat_cash": 15, "ad_cash": 16, "fix_cash": 17,
    "annual_cash": 18, "init_cash": 19, "cash_out": 20,
    "net_cf": 22, "cum_cf": 23,
}
CF_BOLD = {"cash_in", "stock_close", "cash_out", "net_cf", "cum_cf"}
CF_HI = {"cum_cf"}
CF_INDENT = {"pay_purchase", "po_ship", "var_cash", "mat_cum", "mat_cash",
             "ad_cash", "fix_cash", "annual_cash", "init_cash"}


def build_cf(wb, rows):
    ws = wb.create_sheet("CF")
    ws.sheet_properties.tabColor = "B7791F"
    ws.column_dimensions["A"].width = 34
    ws["A1"] = "月次キャッシュフロー（M0=開業準備月〜M18）"
    ws["A1"].font = title_f
    month_header(ws, MONTHS + 2, m_start=0)

    labels = {
        "cash_in": "現金収入（売上−決済手数料）",
        "stock_open": "期首在庫", "purchase": "仕入（発注＝同月前払い）",
        "cogs_use": "原価消費", "stock_close": "期末在庫",
        "pay_purchase": "仕入支払", "po_ship": "仕入送料",
        "var_cash": "発送送料+破損（現金）", "mat_cum": "資材消費 累計（参考）",
        "mat_cash": "資材購入（初期ロット消化後）", "ad_cash": "広告費",
        "fix_cash": "撮影ランニング（現金）", "annual_cash": "年払い（Shopify+ドメイン）",
        "init_cash": "初期投資（撮影・資材・開業）", "cash_out": "現金支出 合計",
        "net_cf": "月次キャッシュフロー", "cum_cf": "累計キャッシュフロー",
    }
    for r, label in CF_SECTIONS.items():
        sec_row(ws, r, MONTHS + 2, label)
    for key, r in CF_ROWS.items():
        name = ("  " + labels[key]) if key in CF_INDENT else labels[key]
        ws.cell(r, 1, name).font = bold_f if key in CF_BOLD else lbl_f
        ws.row_dimensions[r].height = H_DATA

    R = CF_ROWS
    P = PL_ROWS
    for m in range(0, MONTHS + 1):
        c = cf_col(m)
        p = cf_col(m - 1) if m > 0 else None
        pc = pl_col(m) if m >= 1 else None            # 当月の PL 列
        pn = pl_col(m + 1) if m < MONTHS else None    # 翌月の PL 列

        if m == 0:
            f = {
                "cash_in": 0,
                "stock_open": 0,
                "purchase": f"={ref(rows,'inv_initial')}",
                "cogs_use": 0,
                "stock_close": f"={c}{R['stock_open']}+{c}{R['purchase']}-{c}{R['cogs_use']}",
                "pay_purchase": f"={c}{R['purchase']}",
                "po_ship": f"=IF({c}{R['purchase']}>0,{ref(rows,'po_ship')},0)",
                "var_cash": 0, "mat_cum": 0, "mat_cash": 0, "ad_cash": 0, "fix_cash": 0,
                "annual_cash": f"={ref(rows,'shopify_year')}+{ref(rows,'domain_year')}",
                "init_cash": f"={ref(rows,'init_photo')}+{ref(rows,'init_mat')}+{ref(rows,'init_open')}",
            }
        else:
            # 発注 = 翌月予測原価×カバー月数を期末在庫が下回る分だけ補充
            next_cogs = f"PL!{pn}{P['cogs']}" if pn else f"PL!{pc}{P['cogs']}*1.05"
            f = {
                "cash_in": f"=PL!{pc}{P['revenue']}-PL!{pc}{P['pay_fee']}",
                "stock_open": f"={p}{R['stock_close']}",
                "purchase": f"=MAX(0,{next_cogs}*{ref(rows,'stock_cover')}"
                            f"-({c}{R['stock_open']}-PL!{pc}{P['cogs']}))",
                "cogs_use": f"=PL!{pc}{P['cogs']}",
                "stock_close": f"={c}{R['stock_open']}+{c}{R['purchase']}-{c}{R['cogs_use']}",
                "pay_purchase": f"={c}{R['purchase']}",
                "po_ship": f"=IF({c}{R['purchase']}>0,{ref(rows,'po_ship')},0)",
                "var_cash": f"=PL!{pc}{P['ship_cost']}+PL!{pc}{P['loss']}",
                "mat_cum": f"={p}{R['mat_cum']}+PL!{pc}{P['mat_cost']}",
                "mat_cash": f"=MAX(0,{c}{R['mat_cum']}"
                            f"-MAX({ref(rows,'init_mat')},{p}{R['mat_cum']}))",
                "ad_cash": f"=PL!{pc}{P['ad']}",
                "fix_cash": f"={ref(rows,'fix_photo')}",
                # 年払い更新はM12固定（構造依存の条件のため生成時に解決）
                "annual_cash": f"={ref(rows,'shopify_year')}+{ref(rows,'domain_year')}"
                               if m == 12 else 0,
                "init_cash": 0,
            }
        f["cash_out"] = (f"={c}{R['pay_purchase']}+{c}{R['po_ship']}+{c}{R['var_cash']}"
                         f"+{c}{R['mat_cash']}+{c}{R['ad_cash']}+{c}{R['fix_cash']}"
                         f"+{c}{R['annual_cash']}+{c}{R['init_cash']}")
        f["net_cf"] = f"={c}{R['cash_in']}-{c}{R['cash_out']}"
        f["cum_cf"] = f"={c}{R['net_cf']}" if m == 0 else f"={p}{R['cum_cf']}+{c}{R['net_cf']}"

        pl_linked = {"cash_in", "cogs_use", "var_cash", "ad_cash"}
        for key, formula in f.items():
            cell = ws.cell(R[key], m + 2, formula)
            cell.border = thin_b
            cell.font = link_f if key in pl_linked else \
                bold_f if key in CF_BOLD else calc_f
            cell.number_format = JPY
            if key in CF_HI:
                cell.fill = hi_fill
            if key in ("stock_close", "cash_out", "cum_cf"):
                cell.border = med_b
    year_border(ws, cf_col(12), 2, max(CF_ROWS.values()))
    ws.freeze_panes = "B3"
    return ws


# ---------------------------------------------------------------------------
# 「サマリー」シート
# ---------------------------------------------------------------------------
def build_summary(wb, rows):
    ws = wb.create_sheet("サマリー", 0)  # 先頭に配置
    ws.sheet_properties.tabColor = C_PRI
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 52

    ws["A1"] = "Bsharp 財務モデル — サマリー"
    ws["A1"].font = title_f
    ws["A2"] = "全て自動計算（緑字 = 他シート参照）。感度分析は「前提」シートを編集"
    ws["A2"].font = sub_f
    ws.row_dimensions[1].height = H_TITLE
    ws.row_dimensions[2].height = H_DATA

    for i, h in enumerate(["指標", "値", "備考"]):
        cell = ws.cell(3, 1 + i, h)
        cell.font, cell.fill, cell.alignment = hdr_f, hdr_fill, ctr
    ws.row_dimensions[3].height = H_DATA

    m12, m18 = pl_col(12), pl_col(MONTHS)
    pl_op = f"PL!B{PL_ROWS['op']}:{m18}{PL_ROWS['op']}"
    cf_cum = f"CF!B{CF_ROWS['cum_cf']}:{cf_col(MONTHS)}{CF_ROWS['cum_cf']}"
    cf_cum_m1 = f"CF!C{CF_ROWS['cum_cf']}:{cf_col(MONTHS)}{CF_ROWS['cum_cf']}"
    items = [
        ("M12 月商", f"=PL!{m12}{PL_ROWS['revenue']}", JPY, "ゴール: ¥300,000"),
        ("M12 営業利益", f"=PL!{m12}{PL_ROWS['op']}", JPY, ""),
        ("月次黒字化", f"=MATCH(TRUE,INDEX({pl_op}>0,0),0)", '"M"0', "営業利益が初めて黒字になる月"),
        ("損益分岐 注文数/月",
         f"=({ref(rows,'fix_total')}+{ref(rows,'ad_budget')})/{ref(rows,'contrib_per_order')}",
         NUM1, "広告込み固定費 ÷ 貢献利益/注文"),
        ("必要自己資金（資金の谷）", f"=-MIN({cf_cum})", JPY, "累計CFの最大マイナス幅"),
        ("資金の谷の月", f"=MATCH(MIN({cf_cum}),{cf_cum},0)-1", '"M"0', "M0起点"),
        ("投資回収（累計CF黒字化）",
         f"=IFERROR(MATCH(TRUE,INDEX({cf_cum_m1}>=0,0),0),\"18ヶ月超\")", '"M"0', ""),
        ("18ヶ月 累計営業利益", f"=PL!{m18}{PL_ROWS['cum_op']}", JPY, ""),
        ("貢献利益/注文", f"={ref(rows,'contrib_per_order')}", JPY, "前提シートの計算値"),
    ]
    for i, (label, formula, fmt, note) in enumerate(items):
        r = 4 + i
        ws.cell(r, 1, label).font = lbl_f
        cell = ws.cell(r, 2, formula)
        cell.font, cell.number_format = link_f, fmt
        cell.fill, cell.border = hi_fill, thin_b
        ws.cell(r, 3, note).font = sub_f
        ws.row_dimensions[r].height = H_DATA

    r = 4 + len(items) + 1
    notes = [
        "使い方: 「前提」シートの青字セルを変えると全シートが再計算される。",
        "シナリオ係数 0.6 / 1.0 / 1.4 で保守 / 標準 / 強気を切替。",
        "構造の変更（行追加・ロジック変更）は docs/finance/build_financial_model.py を編集して再生成する。",
    ]
    for i, t in enumerate(notes):
        ws.cell(r + i, 1, t).font = sub_f


# ---------------------------------------------------------------------------
# Python プレビュー: Excel 式のミラー計算（式の妥当性チェック用）
# ---------------------------------------------------------------------------
def preview():
    p = {row[0]: row[2] for row in PARAMS if row[2] is not None}
    shares = [b[1] for b in BUCKETS]
    vals = [b[2] for b in BUCKETS]
    ships = [b[3] for b in BUCKETS]
    mats = [b[4] for b in BUCKETS]
    flags = [1 if v >= p["free_line"] else 0 for v in vals]
    aov = sum(s * v for s, v in zip(shares, vals))
    ship_income = sum(s * (1 - fl) for s, fl in zip(shares, flags)) * p["ship_charge"]
    ship_cost = sum(s * v for s, v in zip(shares, ships))
    mat_cost = sum(s * v for s, v in zip(shares, mats))
    fix_total = p["fix_shopify"] + p["fix_domain"] + p["fix_photo"]

    months = []
    cum_op = 0.0
    stock = p["inv_initial"]
    cum_cf = -(p["inv_initial"] + p["po_ship"] + p["shopify_year"] + p["domain_year"]
               + p["init_photo"] + p["init_mat"] + p["init_open"])
    trough, trough_m = cum_cf, 0
    payback = None
    mat_cum_prev = 0.0
    cogs_list = []

    # 1パス目: 注文と原価を先に求める（発注式が翌月原価を参照するため）
    tmp = []
    cc = 0.0
    for m in range(1, MONTHS + 1):
        sess = round((p["sess_init"] + p["sess_grow"] * (m - 1)) * p["scenario"])
        cvr = p["cvr_base"] + p["cvr_slope"] * max(0, m - 3)
        boost = p["boost_m1"] if m == 1 else p["boost_m2"] if m == 2 else 0
        rt = p["ad_budget"] / p["rt_cpc"] * p["rt_cvr"] if m >= p["ad_start"] else 0
        new = sess * cvr + boost + rt
        rep = cc * p["repeat_rate"] if m > 1 else 0
        cc += new
        orders = new + rep
        tmp.append((sess, cvr, orders))
        cogs_list.append(orders * aov * p["kakeritsu"])

    for m in range(1, MONTHS + 1):
        sess, cvr, orders = tmp[m - 1]
        product_rev = orders * aov
        revenue = product_rev + orders * ship_income
        cogs = cogs_list[m - 1]
        next_cogs = cogs_list[m] if m < MONTHS else cogs_list[-1] * 1.05
        purchase = max(0.0, next_cogs * p["stock_cover"] - (stock - cogs))
        stock = stock + purchase - cogs
        po_ship = p["po_ship"] if purchase > 0 else 0
        pay_fee = revenue * p["pay_rate"]
        contrib = (revenue - cogs - pay_fee - orders * ship_cost - orders * mat_cost
                   - product_rev * p["loss_rate"] - po_ship)
        ad = p["ad_budget"] if m >= p["ad_start"] else 0
        op = contrib - ad - fix_total
        cum_op += op

        mat_cum = mat_cum_prev + orders * mat_cost
        mat_cash = max(0.0, mat_cum - max(p["init_mat"], mat_cum_prev))
        mat_cum_prev = mat_cum
        cash_in = revenue - pay_fee
        cash_out = (purchase + po_ship + orders * ship_cost + product_rev * p["loss_rate"]
                    + mat_cash + ad + p["fix_photo"]
                    + (p["shopify_year"] + p["domain_year"] if m == 12 else 0))
        cum_cf += cash_in - cash_out
        if cum_cf < trough:
            trough, trough_m = cum_cf, m
        if payback is None and cum_cf >= 0:
            payback = m
        months.append((m, sess, orders, revenue, op, cum_cf, stock))

    print(f"{'月':>4} {'セッション':>7} {'注文':>6} {'月商':>9} {'営業利益':>9} {'累計CF':>10} {'在庫':>9}")
    for m, sess, orders, revenue, op, cf, stk in months:
        if m in (1, 2, 3, 6, 9, 12, 15, 18):
            print(f"M{m:<3} {sess:>7,} {orders:>6.1f} {revenue:>9,.0f} {op:>9,.0f} {cf:>10,.0f} {stk:>9,.0f}")
    print(f"\n資金の谷: ¥{-trough:,.0f}（M{trough_m}） / 累計CF黒字化: "
          f"{'M' + str(payback) if payback else '18ヶ月超'} / 18ヶ月累計営業利益: ¥{cum_op:,.0f}")


# ---------------------------------------------------------------------------
def main():
    out = Path(__file__).resolve().parent / "financial-model.xlsx"
    wb = Workbook()
    rows = build_assumptions(wb)
    build_pl(wb, rows)
    build_cf(wb, rows)
    build_summary(wb, rows)
    wb.save(out)
    print(f"生成完了: {out}\n")
    print("=== プレビュー（Excel式のミラー計算・標準シナリオ） ===")
    preview()


if __name__ == "__main__":
    main()
